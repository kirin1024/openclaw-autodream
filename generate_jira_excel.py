#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from collections import defaultdict

# 从Jira页面提取的数据（已过滤：排除"无需前端"）
# 字段: [需求类型(优先级), 需求标题, 跟进人, 本期占用估时(d), 状态, 发布版本, Jira链接]
data = [
    # FLOWMORE-13283 - 陈燕跟进
    ["P1", "前端pc-福贸注册流程接入制裁IP拦截", "陈燕", 2, "待解决", "", "http://jira.pingpongx.com/browse/FLOWMORE-13283"],
    # FLOWMORE-13281 - 周腾跟进
    ["P1", "前端-小程序代言人佣金结算优化", "周腾", 0.5, "待解决", "", "http://jira.pingpongx.com/browse/FLOWMORE-13281"],
    # FLOWMORE-13280 - 唐雪跟进
    ["P1", "前端-移动端-【账单收款】账单支付的地区取值优化", "唐雪", 1, "待解决", "", "http://jira.pingpongx.com/browse/FLOWMORE-13280"],
    # FLOWMORE-13277 - 周腾跟进 (前端估时1)
    ["P1", "BNI入账补充付款人信息", "周腾", 1, "待整理", "", "http://jira.pingpongx.com/browse/FLOWMORE-13277"],
    # FLOWMORE-13276 - 陈燕跟进
    ["P1", "短信验证码增加登陆场景", "陈燕", 1, "待整理", "", "http://jira.pingpongx.com/browse/FLOWMORE-13276"],
    # FLOWMORE-13272 - 袁梦苏跟进
    ["P0", "代言人佣金结算优化", "袁梦苏", "", "待整理", "", "http://jira.pingpongx.com/browse/FLOWMORE-13272"],
    # FLOWMORE-13265 - 周腾跟进
    ["P1", "【账单收款】部分支付的账单允许关闭", "周腾", 0.5, "待整理", "", "http://jira.pingpongx.com/browse/FLOWMORE-13265"],
    # FLOWMORE-13264 - 陈燕跟进
    ["P1", "前端小程序、h5-福贸注册流程接入制裁IP拦截", "陈燕", 1, "待解决", "", "http://jira.pingpongx.com/browse/FLOWMORE-13264"],
    # FLOWMORE-13262 - 陈燕跟进
    ["P0", "福贸注册流程接入制裁IP拦截", "陈燕", "", "待开发", "", "http://jira.pingpongx.com/browse/FLOWMORE-13262"],
    # FLOWMORE-13235 - 陈燕跟进
    ["P1", "【账单收款】账单支付的地区取值优化", "陈燕", 1, "待整理", "", "http://jira.pingpongx.com/browse/FLOWMORE-13235"],
    # FLOWMORE-13050 - 周腾跟进
    ["P1", "JPY出金取消下拉选择银行", "周腾", 0.5, "待整理", "", "http://jira.pingpongx.com/browse/FLOWMORE-13050"],
    # FLOWMORE-12845 - 周腾跟进
    ["P1", "开放境外地区-美金薪酬账户绑定", "周腾", 1, "待整理", "", "http://jira.pingpongx.com/browse/FLOWMORE-12845"],
    # FLOWMORE-12514 - 周腾跟进
    ["P1", "【创建合同】上传合同附件后支持OCR识别", "周腾", 1, "待整理", "", "http://jira.pingpongx.com/browse/FLOWMORE-12514"],
    # FLOWMORE-12309 - 周腾跟进
    ["P1", "信用证工商银行接入", "周腾", 2, "待整理", "", "http://jira.pingpongx.com/browse/FLOWMORE-12309"],
]

# 按跟进人排序
data.sort(key=lambda x: x[2])

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "Jira需求"

# 定义表头
headers = ["需求类型（优先级）", "需求标题", "跟进人", "状态", "测试安排", "本期占用估时(d)", "发布版本", "开发版本", "备注"]

# 写入表头
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 写入数据
for row_idx, row_data in enumerate(data, 2):
    # row_data: [优先级, 标题, 跟进人, 估时, 状态, 发布版本, 链接]
    # Excel列: 需求类型, 需求标题, 跟进人, 状态, 测试安排, 本期占用估时, 发布版本, 开发版本, 备注
    ws.cell(row=row_idx, column=1, value=row_data[0])  # 需求类型
    ws.cell(row=row_idx, column=2, value=row_data[1])  # 需求标题
    ws.cell(row=row_idx, column=3, value=row_data[2])  # 跟进人
    ws.cell(row=row_idx, column=4, value=row_data[4])  # 状态
    ws.cell(row=row_idx, column=5, value="")  # 测试安排（留空）
    ws.cell(row=row_idx, column=6, value=row_data[3])  # 本期占用估时
    ws.cell(row=row_idx, column=7, value=row_data[5])  # 发布版本
    ws.cell(row=row_idx, column=8, value="")  # 开发版本（留空）
    ws.cell(row=row_idx, column=9, value=row_data[6])  # 备注(Jira链接)

# 调整列宽
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 12
ws.column_dimensions['I'].width = 45

# 保存文件
output_path = "/Users/chenyan/.openclaw/workspace/Jira需求_26.0323_筛选后.xlsx"
wb.save(output_path)

print(f"Excel文件已保存到: {output_path}")

# 统计
print(f"\n共筛选出 {len(data)} 条需求")

# 按跟进人统计
stats = defaultdict(int)
for row in data:
    stats[row[2]] += 1

print("\n按跟进人统计:")
for person, count in sorted(stats.items(), key=lambda x: -x[1]):
    print(f"  {person}: {count}条")
